#!/usr/bin/env python3
"""
Recovery Script for Corrupted Sample CSV

One-time script to recover the corrupted sample_full_annotated_v2.csv file.
The corruption was caused by opening a UTF-8 CSV in Excel, which:
    1. Converted Swedish characters (ä, ö, å) to garbled text
    2. Inserted hidden control characters
    3. Misaligned some cells

This script:
    1. Loads the corrupted CSV with latin-1 encoding
    2. Fixes doc_ids using a Swedish character mapping
    3. Strips hidden control characters
    4. Joins to source parquet to recover correct sentence_text
    5. Preserves annotations (mechanism columns, coder_notes, split)
    6. Outputs clean Excel file ready for BERT training

Usage:
    python recover_corrupted_sample.py \\
        --input results/00_data_preparation/sampling/sample_full_annotated_v2.csv \\
        --corpus data/processed/bert_corpus.parquet \\
        --output results/00_data_preparation/sampling/sample_full_recovered.xlsx

Author: Swedish Risk Analysis Text-as-Data Project
Date: 2026-03-03
"""

import argparse
import re
import sys
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows


# =============================================================================
# CONFIGURATION
# =============================================================================

# Mapping from corrupted Swedish character patterns to correct characters
SWEDISH_CHAR_MAP = {
    'A_tvidaberg': 'Åtvidaberg',
    'A_nge': 'Ånge',
    'A_re': 'Åre',
    'Ga_llivare': 'Gällivare',
    'Go_tene': 'Götene',
    'Ho_gans': 'Höganäs',
    'Kungso_r': 'Kungsör',
    'La_n': 'Län',
    'Malmo_': 'Malmö',
    'Norsjo_': 'Norsjö',
    'O_rebro': 'Örebro',
    'O_sterker': 'Österåker',
    'O_stersund': 'Östersund',
    'Sa_ffle': 'Säffle',
    'Sko_vde': 'Skövde',
    'So_dermanland': 'Södermanland',
    'So_der': 'Söder',
    'Stro_mstad': 'Strömstad',
    'Stro_msund': 'Strömsund',
    'Timra_': 'Timrå',
    'Umea_': 'Umeå',
    'Va_rmd': 'Värmdö',
    'Va_rmland': 'Värmland',
    'Va_stervik': 'Västervik',
    'Va_ster': 'Väster',
    'Va_xj': 'Växjö',
    'Va_xjo_': 'Växjö',
    'O_ster': 'Öster',
}

# Columns to preserve from annotations
ANNOTATION_COLUMNS = [
    'mechanism_legitimacy',
    'mechanism_functional',
    'mechanism_equivalence',
    'mechanism_complexity',
    'coder_notes',
]

# Metadata columns to preserve
METADATA_COLUMNS = [
    'sample_id',
    'split',
]


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def fix_doc_id(doc_id: str) -> str:
    """
    Fix corrupted doc_id by applying Swedish character mapping.

    Parameters
    ----------
    doc_id : str
        Corrupted doc_id string.

    Returns
    -------
    str
        Fixed doc_id with correct Swedish characters.
    """
    if pd.isna(doc_id):
        return doc_id

    # Remove non-printable/control characters
    doc_id = ''.join(c for c in doc_id if c.isprintable() or c in ' \n\t')

    # Remove trailing garbage after .pdf/.PDF
    doc_id = re.sub(r'(\.pdf|\.PDF)[a-z ]*$', r'\1', doc_id, flags=re.IGNORECASE)
    doc_id = doc_id.strip()

    # Apply Swedish character mappings
    for bad, good in SWEDISH_CHAR_MAP.items():
        doc_id = doc_id.replace(bad, good)

    # Normalize Unicode to NFC form
    return unicodedata.normalize('NFC', doc_id)


def save_excel_formatted(df: pd.DataFrame, path: Path, text_column: str = 'sentence_text') -> None:
    """
    Save DataFrame to Excel with formatting for hand-coding.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame to save.
    path : Path
        Output path (.xlsx).
    text_column : str
        Name of the text column to wrap.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample"

    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True)
            col_name = df.columns[c_idx - 1] if c_idx <= len(df.columns) else None
            if col_name == text_column:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Adjust column widths
    for col_idx, column in enumerate(df.columns, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        if column == text_column:
            ws.column_dimensions[col_letter].width = 80
        else:
            max_length = max(
                len(str(column)),
                df[column].astype(str).str.len().max() if len(df) > 0 else 0
            )
            ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

    # Set row height for text rows
    for row_idx in range(2, len(df) + 2):
        ws.row_dimensions[row_idx].height = 60

    wb.save(path)


# =============================================================================
# MAIN RECOVERY LOGIC
# =============================================================================

def recover_sample(
    input_path: Path,
    corpus_path: Path,
    output_path: Path,
    verbose: bool = False,
) -> dict:
    """
    Recover corrupted sample by joining back to source corpus.

    Parameters
    ----------
    input_path : Path
        Path to corrupted CSV.
    corpus_path : Path
        Path to source parquet (bert_corpus.parquet or bert_corpus_filtered.parquet).
    output_path : Path
        Path for output Excel file.
    verbose : bool
        Print detailed progress.

    Returns
    -------
    dict
        Recovery statistics.
    """
    print("=" * 60)
    print("RECOVERING CORRUPTED SAMPLE")
    print("=" * 60)

    # Load corrupted CSV
    print(f"\nLoading corrupted CSV: {input_path}")
    corrupted = pd.read_csv(input_path, sep=';', encoding='latin-1')
    print(f"  Rows: {len(corrupted)}")

    # Load source corpus
    print(f"\nLoading source corpus: {corpus_path}")
    corpus = pd.read_parquet(corpus_path)
    print(f"  Rows: {len(corpus)}")

    # Normalize corpus doc_ids
    corpus['doc_id_norm'] = corpus['doc_id'].apply(
        lambda x: unicodedata.normalize('NFC', x) if pd.notna(x) else x
    )
    corpus['key'] = corpus['doc_id_norm'] + '_' + corpus['sentence_id'].astype(str)

    # Fix corrupted doc_ids
    print("\nFixing corrupted doc_ids...")
    corrupted['doc_id_fixed'] = corrupted['doc_id'].apply(fix_doc_id)
    corrupted['key'] = corrupted['doc_id_fixed'] + '_' + corrupted['sentence_id'].astype(str)

    # Match to corpus
    print("\nMatching to source corpus...")
    matched_mask = corrupted['key'].isin(corpus['key'])
    n_matched = matched_mask.sum()
    n_unmatched = (~matched_mask).sum()
    print(f"  Matched: {n_matched}")
    print(f"  Unmatched: {n_unmatched}")

    if n_unmatched > 0:
        print("\n  Unmatched doc_ids:")
        for doc in corrupted[~matched_mask]['doc_id_fixed'].unique()[:10]:
            print(f"    {doc}")

    # Create key-indexed corpus for joining
    corpus_indexed = corpus.set_index('key')

    # Build recovered dataframe
    print("\nRecovering correct sentence_text...")
    recovered_rows = []

    for _, row in corrupted.iterrows():
        key = row['key']

        if key in corpus_indexed.index:
            # Get correct data from corpus
            corpus_row = corpus_indexed.loc[key]

            # Build recovered row
            recovered = {
                'sample_id': row['sample_id'],
                'doc_id': corpus_row['doc_id'],  # Use correct doc_id from corpus
                'entity': corpus_row.get('municipality', row.get('entity', '')),
                'actor_type': corpus_row['actor_type'],
                'year': corpus_row['year'],
                'wave': row.get('wave', ''),
                'sentence_id': corpus_row['sentence_id'],
                'paragraph_id': corpus_row.get('paragraph_id', row.get('paragraph_id', '')),
                'sentence_text': corpus_row['sentence_text'],  # Correct text!
                'word_count': corpus_row.get('word_count', row.get('word_count', '')),
                'split': row['split'],
            }

            # Preserve annotations
            for col in ANNOTATION_COLUMNS:
                if col in row.index:
                    recovered[col] = row[col]

            recovered_rows.append(recovered)
        else:
            # Row couldn't be matched - include with warning
            if verbose:
                print(f"  WARNING: Could not match {row['sample_id']}: {row['doc_id']}")

    recovered_df = pd.DataFrame(recovered_rows)

    # Reorder columns
    column_order = (
        ['sample_id', 'doc_id', 'entity', 'actor_type', 'year', 'wave',
         'sentence_id', 'paragraph_id', 'sentence_text', 'word_count', 'split']
        + ANNOTATION_COLUMNS
    )
    recovered_df = recovered_df[[c for c in column_order if c in recovered_df.columns]]

    # Save recovered Excel
    print(f"\nSaving recovered Excel: {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    save_excel_formatted(recovered_df, output_path)

    # Statistics
    stats = {
        'input_rows': len(corrupted),
        'matched_rows': n_matched,
        'unmatched_rows': n_unmatched,
        'output_rows': len(recovered_df),
        'annotations_preserved': {
            col: int(recovered_df[col].notna().sum())
            for col in ANNOTATION_COLUMNS
            if col in recovered_df.columns
        },
    }

    print("\n" + "=" * 60)
    print("RECOVERY COMPLETE")
    print("=" * 60)
    print(f"\nInput rows: {stats['input_rows']}")
    print(f"Matched: {stats['matched_rows']}")
    print(f"Unmatched (dropped): {stats['unmatched_rows']}")
    print(f"Output rows: {stats['output_rows']}")
    print(f"\nAnnotations preserved:")
    for col, count in stats['annotations_preserved'].items():
        print(f"  {col}: {count}")
    print(f"\nOutput: {output_path}")

    return stats


# =============================================================================
# CLI
# =============================================================================

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Recover corrupted sample CSV file",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    parser.add_argument(
        '--input',
        type=Path,
        default=Path('results/00_data_preparation/sampling/sample_full_annotated_v2.csv'),
        help='Path to corrupted CSV file',
    )
    parser.add_argument(
        '--corpus',
        type=Path,
        default=Path('data/processed/bert_corpus.parquet'),
        help='Path to source corpus parquet (bert_corpus.parquet has all docs)',
    )
    parser.add_argument(
        '--output',
        type=Path,
        default=Path('results/00_data_preparation/sampling/sample_full_recovered.xlsx'),
        help='Path for output Excel file',
    )
    parser.add_argument(
        '--verbose',
        action='store_true',
        help='Print detailed progress',
    )

    args = parser.parse_args()

    if not args.input.exists():
        print(f"Error: Input file not found: {args.input}", file=sys.stderr)
        return 1

    if not args.corpus.exists():
        print(f"Error: Corpus file not found: {args.corpus}", file=sys.stderr)
        return 1

    try:
        recover_sample(
            input_path=args.input,
            corpus_path=args.corpus,
            output_path=args.output,
            verbose=args.verbose,
        )
        return 0
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1


if __name__ == '__main__':
    sys.exit(main())
